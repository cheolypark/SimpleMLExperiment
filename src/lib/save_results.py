from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class SaveResults():
    def __init__(self, filename):
        self.wb = Workbook()
        self._filename = filename
        self._colCount = 1

    def insert(self, col_name=None, re_list=None, wstitle="Prediction Results"):
        ws1 = self.wb.active
        ws1.title = wstitle

        ws1.cell(row=1, column=self._colCount, value=col_name)
        for i, y in enumerate(re_list, start=2):
            ws1.cell(row=i, column=self._colCount, value=y)

        self._colCount += 1

    def put(self, df, wstitle="results"):
        ws = self.wb.active
        ws.title = wstitle
        rows = dataframe_to_rows(df)

        row_count = 0
        for r_idx, row in enumerate(rows, 1):
            if len(row) > 1:
                row_count += 1

                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=row_count, column=c_idx, value=value)

    def save(self):
        self.wb.save(filename=self._filename)
